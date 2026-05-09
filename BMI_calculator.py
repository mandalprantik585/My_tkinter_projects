from tkinter import *
from tkinter import font, messagebox
from openpyxl import Workbook, load_workbook
from datetime import date
import os

r = Tk()
r.title("BMI calculator")
file_name = "BMI_data.xlsx"
r.geometry("700x500")

def Preview():
    os.startfile(file_name)

def calculate():
    try:
        name = name_entry.get()
        age = int(age_entry.get())
        hight = int(hight_entry.get())
        weight = int(weight_entry.get())
        BMI = round(weight/((hight/100) ** 2), 2)
        BMI_entry.delete(0, END)
        status = ""
        if BMI >= 25:
            status = "Over weight"
        elif BMI < 18:
            status = "Under weight"
        else:
            status = "Normal"
        BMI_entry.insert(END, f"{BMI}, {status}")
        if os.path.exists(file_name):
            wb = load_workbook(file_name)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Date", "Name", "Age", "Height(in cm)", "Weight(in Kg)", "BMI", "Status"])
        ws.append([str(date.today()), name, str(age), str(hight), str(weight), str(BMI), status])
        wb.save(file_name)
        messagebox.showinfo("Success", "Data have been saved successfully!")

        preview_button = Button(r, text = "Preview", font = font.Font(size = 25), width = 10, border = 5, command = Preview)
        preview_button.grid(sticky = W, row = 6, column = 1, padx = 100, pady = 10)     
    except:
        messagebox.showerror("Error", "Something went wrong")          

Label(r, text = "Enter your name : ", font = font.Font(size = 17)).grid(sticky = W, row = 0, padx = 10, pady = 10)
name_entry = Entry(r, font = font.Font(size = 17), width = 25, border = 3)
name_entry.grid(row = 0, padx = 10, column = 1)

Label(r, text = "Enter your age : ", font = font.Font(size = 17)).grid(sticky = W, row = 1, padx = 10, pady = 10)
age_entry = Entry(r, font = font.Font(size = 17), width = 25, border = 3)
age_entry.grid(row = 1, padx = 10, column = 1)

Label(r, text = "Enter your weight(in Kg) : ", font = font.Font(size = 17)).grid(sticky = W, row = 2, padx = 10, pady = 10)
weight_entry = Entry(r, font = font.Font(size = 17), width = 25, border = 3)
weight_entry.grid(row = 2, padx = 10, column = 1)

Label(r, text = "Enter your hight(in cm) : ", font = font.Font(size = 17)).grid(sticky = W, row = 3, padx = 10, pady = 10)
hight_entry = Entry(r, font = font.Font(size = 17), width = 25, border = 3)
hight_entry.grid(row = 3, padx = 10, column = 1)

go_button = Button(r, text = "GO", font = font.Font(size = 25), width = 5, border = 5, command = calculate)
go_button.grid(sticky = W, row = 4, column = 1, padx = 100, pady = 10)

Label(r, text = "BMI : ", font = font.Font(size = 25)).grid(sticky = W, row = 5, padx = 10, pady = 20)
BMI_entry = Entry(r, font = font.Font(size = 25), width = 20, border = 3)
BMI_entry.grid(sticky = W, row = 5, column = 1, padx = 10, pady = 10)

r.mainloop()